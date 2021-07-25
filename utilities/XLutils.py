import openpyxl
def getrow(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook['Sheet1']
    return (sheet.max_row)
def getcolumn(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook['Sheet1']
    return (sheet.max_column)
def readdata(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook['Sheet1']
    return sheet.cell(rownum,columnno).value

def writedata(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook['Sheet1']
    sheet.cell(rownum,columnno).value=data

